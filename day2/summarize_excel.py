import openpyxl
import json

def summarize_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Get columns
        columns = [cell.value for cell in sheet[1]]
        
        # Get data
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
            data.append(dict(zip(columns, row)))
            
        summary = {
            "columns": columns,
            "total_rows": sheet.max_row - 1,
            "sample_data": data
        }
        
        print(json.dumps(summary, indent=2))
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    summarize_excel('floodarchive.xlsx')
